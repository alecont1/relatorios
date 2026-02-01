"""
Excel parsing service for template import.

Parses Excel files with expected format:
| Section | Script Step | Result Type | Step Result Values |

Result Type: "Drop Down" or "Text"
Step Result Values: comma or slash separated options for dropdowns (e.g., "Yes/No/NA")
"""

from io import BytesIO
from dataclasses import dataclass, field
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


@dataclass
class ParsedField:
    label: str
    field_type: str  # "dropdown" or "text"
    options: list[str] | None
    order: int


@dataclass
class ParsedSection:
    name: str
    fields: list[ParsedField] = field(default_factory=list)
    order: int = 0


@dataclass
class ParseResult:
    valid: bool
    sections: list[ParsedSection] | None = None
    errors: list[str] | None = None
    summary: dict | None = None


def parse_template_excel(file_content: bytes) -> ParseResult:
    """
    Parse Excel file and return structured template data or errors.

    Expected columns (row 1 is header):
    A: Section - section name (required)
    B: Script Step - question/field label (required)
    C: Result Type - "Drop Down" or "Text" (required)
    D: Step Result Values - options for dropdown, comma or slash separated

    Args:
        file_content: Raw bytes of the uploaded Excel file

    Returns:
        ParseResult with either sections or errors
    """
    errors: list[str] = []

    # Try to load workbook
    try:
        wb = load_workbook(filename=BytesIO(file_content), read_only=True, data_only=True)
    except InvalidFileException:
        return ParseResult(valid=False, errors=["Arquivo Excel invalido ou corrompido"])
    except Exception as e:
        return ParseResult(valid=False, errors=[f"Erro ao abrir arquivo: {str(e)}"])

    sheet = wb.active
    if sheet is None:
        return ParseResult(valid=False, errors=["Arquivo Excel sem planilha ativa"])

    # Validate header row
    header_row = [cell for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    expected_headers = ["Section", "Script Step", "Result Type", "Step Result Values"]

    if len(header_row) < 4:
        errors.append(f"Cabecalho incompleto. Esperado: {', '.join(expected_headers)}")

    # Process data rows
    sections_dict: dict[str, ParsedSection] = {}
    field_order = 0
    section_order = 0

    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        # Skip completely empty rows
        if all(cell is None or str(cell).strip() == "" for cell in row[:4]):
            continue

        # Extract values (handle None gracefully)
        section_name = str(row[0]).strip() if row[0] else ""
        script_step = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        result_type = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        step_values = str(row[3]).strip() if len(row) > 3 and row[3] else ""

        # Validate required fields
        if not section_name:
            errors.append(f"Linha {row_num}: Coluna 'Section' e obrigatoria")

        if not script_step:
            errors.append(f"Linha {row_num}: Coluna 'Script Step' e obrigatoria")

        if not result_type:
            errors.append(f"Linha {row_num}: Coluna 'Result Type' e obrigatoria")
        elif result_type.lower() not in ["drop down", "dropdown", "text"]:
            errors.append(f"Linha {row_num}: 'Result Type' invalido '{result_type}'. Use 'Drop Down' ou 'Text'")

        # Validate dropdown options
        field_type = "dropdown" if result_type.lower() in ["drop down", "dropdown"] else "text"
        options: list[str] | None = None

        if field_type == "dropdown":
            if not step_values:
                errors.append(f"Linha {row_num}: 'Step Result Values' e obrigatorio para tipo 'Drop Down'")
            else:
                # Parse options (handle both / and , separators)
                if "/" in step_values:
                    options = [opt.strip() for opt in step_values.split("/") if opt.strip()]
                elif "," in step_values:
                    options = [opt.strip() for opt in step_values.split(",") if opt.strip()]
                else:
                    options = [step_values.strip()]

                if len(options) < 2:
                    errors.append(f"Linha {row_num}: 'Drop Down' precisa de pelo menos 2 opcoes")

        # If no errors for this row, add to structure
        if section_name and script_step and result_type:
            if section_name not in sections_dict:
                sections_dict[section_name] = ParsedSection(
                    name=section_name,
                    order=section_order
                )
                section_order += 1

            sections_dict[section_name].fields.append(ParsedField(
                label=script_step,
                field_type=field_type,
                options=options,
                order=field_order
            ))
            field_order += 1

    wb.close()

    # Return result
    if errors:
        return ParseResult(valid=False, errors=errors)

    sections_list = list(sections_dict.values())

    if not sections_list:
        return ParseResult(valid=False, errors=["Nenhum dado encontrado no arquivo Excel"])

    # Calculate summary
    total_fields = sum(len(s.fields) for s in sections_list)

    return ParseResult(
        valid=True,
        sections=sections_list,
        summary={
            "section_count": len(sections_list),
            "field_count": total_fields
        }
    )
